"""
Este módulo contém funções para realizar consulta a banco de dados,
tratamento dos dados da consulta e exportação dos dados tratados para Excel.

Autor: Marcos Trömel
E-mail: marcos.tromel@gmail.com
Data de criação: 21/08/2024

Licença: MIT License

Copyright (c) 2024 Marcos Trömel

Permissão é concedida, gratuitamente, a qualquer pessoa que obtenha uma cópia
deste software e de seus arquivos de documentação associados (o "Software"),
para lidar com o Software sem restrições, incluindo, sem limitação, os direitos
de usar, copiar, modificar, mesclar, publicar, distribuir, sublicenciar, e/ou
vender cópias do Software, e de permitir que pessoas a quem o Software seja
fornecido o façam, sob as seguintes condições:

O aviso de copyright acima e este aviso de permissão devem ser incluídos em
todas as cópias ou partes substanciais do Software.

O SOFTWARE É FORNECIDO "TAL COMO ESTÁ", SEM GARANTIA DE QUALQUER TIPO, EXPRESSA
OU IMPLÍCITA, INCLUINDO, MAS NÃO SE LIMITANDO A, GARANTIAS DE COMERCIALIZAÇÃO E
ADEQUAÇÃO A UMA FINALIDADE ESPECÍFICA. EM NENHUM CASO OS AUTORES OU DETENTORES
DE DIREITOS AUTORAIS SERÃO RESPONSÁVEIS POR QUALQUER REIVINDICAÇÃO, DANOS OU
OUTRA RESPONSABILIDADE, SEJA EM UMA AÇÃO CONTRATUAL, DELITO OU OUTRA,
DECORRENTES DE, OU EM CONEXÃO COM O SOFTWARE OU O USO OU OUTRAS NEGOCIAÇÕES NO
SOFTWARE.
"""


from datetime import datetime
from pytz import timezone
from pathlib import Path
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.worksheet.cell_range import CellRange
import pyodbc

def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")

    rng = CellRange(cell_range)
    for row in ws.iter_rows(min_row=rng.min_row,
                            min_col=rng.min_col,
                            max_col=rng.max_col,
                            max_row=rng.max_row):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def set_font_bold(ws, cell_range):
    rng = CellRange(cell_range)
    for row in ws.iter_rows(min_row=rng.min_row,
                            min_col=rng.min_col,
                            max_col=rng.max_col,
                            max_row=rng.max_row):
        for cell in row:
            cell.font = Font(b=True)

def set_font_size(ws, cell_range):
    rng = CellRange(cell_range)
    for row in ws.iter_rows(min_row=rng.min_row,
                            min_col=rng.min_col,
                            max_col=rng.max_col,
                            max_row=rng.max_row):
        for cell in row:
            cell.font = Font(size=11)

def set_align_center(ws, cell_range):
    rng = CellRange(cell_range)
    for row in ws.iter_rows(min_row=rng.min_row,
                            min_col=rng.min_col,
                            max_col=rng.max_col,
                            max_row=rng.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

def set_style_currency(ws, cell_range):
    rng = CellRange(cell_range)
    for row in ws.iter_rows(min_row=rng.min_row,
                            min_col=rng.min_col,
                            max_col=rng.max_col,
                            max_row=rng.max_row):
        for cell in row:
            cell.style = 'Currency'

def set_style_color_yellow(ws, cell_range):
    rng = CellRange(cell_range)
    for row in ws.iter_rows(min_row=rng.min_row,
                            min_col=rng.min_col,
                            max_col=rng.max_col,
                            max_row=rng.max_row):
        for cell in row:
            cell.fill = PatternFill('solid', fgColor='00FFFF99')

TABLE_INVENTSUM = 'dbo.inventsum'
TABLE_INVENTDIM = 'dbo.inventdim'
TABLE_INVENTTABLE = 'dbo.inventtable'
TABLE_INVENTITEMPRICE = 'dbo.inventitemprice'
ROOT_FOLDER = Path(__file__).parent 
WB_REPORT_PATH = ROOT_FOLDER / 'Relatório de Estoque.xlsx'

dep01, dep02, dep03, dep04, dep11, dep12, dep13, dep14, dep15 = 0, 0, 0, 0, 0, 0, 0, 0, 0
dep19, dep20, dep21, dep96, dep97, dep98, dep00 = 0, 0, 0, 0, 0, 0, 0
sum_total = 0
day = datetime.now(timezone('America/Sao_Paulo'))
fmt = '%d/%m/%Y'
report_day = datetime.strftime(day, fmt)

data = [
    ['Nº do item', 'Nome do item', 'Depósito', 'Localização', 'Estoque físico',
      'Unidade de estoque', 'Preço unitário', 'Preço total'],
]

invent_sum = [
    ['Depósito', 'Nome', 'Valor'],
    ['1', 'Almoxarifado', dep01],
    ['2', 'Material usado', dep02],
    ['3', 'Produto acabado', dep03],
    ['4', 'Em produção', dep04],
    ['11', 'Remessa conserto', dep11],
    ['12', 'Remessa garantia', dep12],
    ['13', 'Demonstração Peças', dep13],
    ['14', 'Demonstraçaõ Máquinas', dep14],
    ['15', 'Remessa industrialização', dep15],
    ['19', 'Conserto máquinas terceiros', dep19],
    ['20', 'Assistência técnica', dep20],
    ['21', 'Ferramentas', dep21],
    ['96', 'Descarte assistência técnica', dep96],
    ['97', 'Descarte comercial', dep97],
    ['98', 'Descarte produção', dep98],
    [' ', 'Sem depósito', dep00],
]

# String de conexão com o banco de dados
conn = pyodbc.connect('Driver={SQL Server};'
                      'Server="CHANGE-ME";'
                      'Database="CHANGE-ME";'
                      'UID="CHANGE-ME";'
                      'PWD="CHANGE-ME"')

# Criar um objeto cursor
cursor = conn.cursor()

# Consulta SQL
query = f"SELECT I.ITEMID, IT.ITEMNAME, D.INVENTLOCATIONID, D.WMSLOCATIONID, I.PHYSICALINVENT, P.UNITID, P.PRICE, I.PHYSICALINVENT * P.PRICE AS TOTAL FROM {TABLE_INVENTSUM} AS I LEFT JOIN {TABLE_INVENTDIM} AS D ON I.INVENTDIMID = D.INVENTDIMID LEFT JOIN {TABLE_INVENTTABLE} AS IT ON I.ITEMID = IT.ITEMID LEFT JOIN {TABLE_INVENTITEMPRICE} AS P ON I.ITEMID = P.ITEMID WHERE I.PHYSICALINVENT > 0 AND I.DATAAREAID = 'gsb' AND D.DATAAREAID = 'gsb' AND P.VERSIONID = ' ' AND P.PRICETYPE = '0' AND P.DATAAREAID = 'gsb'"
cursor.execute(query)

# Carregando o arquivo do Excel
wb_report = Workbook()

# Nome para a planilha
sheet_name_report = 'Estoque'
sheet_name_report_sum = 'Inventário'
wb_report.create_sheet(sheet_name_report, 0)
wb_report.create_sheet(sheet_name_report_sum, 1)

# Selecionou a planilha
ws_report: Worksheet = wb_report[sheet_name_report]

# Remover uma planilha
wb_report.remove(wb_report['Sheet'])

for row in cursor:
    item, name, dep, loc, fis_inv, unitid, price, ttl = row
    sum_total += ttl
    data += [[item, name, dep, loc, fis_inv, unitid, price, ttl]]
    if dep == '01':
        dep01 += ttl
    elif dep == '02':
        dep02 += ttl
    elif dep == '03':
        dep03 += ttl
    elif dep == '04':
        dep04 += ttl
    elif dep == '11':
        dep11 += ttl
    elif dep == '12':
        dep12 += ttl
    elif dep == '13':
        dep13 += ttl
    elif dep == '14':
        dep14 += ttl
    elif dep == '15':
        dep15 += ttl
    elif dep == '19':
        dep19 += ttl
    elif dep == '20':
        dep20 += ttl
    elif dep == '21':
        dep21 += ttl
    elif dep == '96':
        dep96 += ttl
    elif dep == '97':
        dep97 += ttl
    elif dep == '98':
        dep98 += ttl
    elif dep == '':
        dep00 += ttl

    sum_contabil = dep00+dep01+dep02+dep03+dep04+dep11+dep12+dep13+dep14+dep15+dep19+dep20+dep21

for line in data:
    ws_report.append(line)


invent_sum = [
    ['Depósito', 'Nome', 'Valor'],
    [' ', 'Sem depósito', dep00],
    ['1', 'Almoxarifado', dep01],
    ['2', 'Material usado', dep02],
    ['3', 'Produto acabado', dep03],
    ['4', 'Em produção', dep04],
    ['11', 'Remessa conserto', dep11],
    ['12', 'Remessa garantia', dep12],
    ['13', 'Demonstração Peças', dep13],
    ['14', 'Demonstraçaõ Máquinas', dep14],
    ['15', 'Remessa industrialização', dep15],
    ['19', 'Conserto máquinas terceiros', dep19],
    ['20', 'Assistência técnica', dep20],
    ['21', 'Ferramentas', dep21],
    ['96', 'Descarte assistência técnica', dep96],
    ['97', 'Descarte comercial', dep97],
    ['98', 'Descarte produção', dep98],
]

ws_report: Worksheet = wb_report[sheet_name_report_sum]

for line in invent_sum:
    ws_report.append(line)

ws_report.insert_rows(1, amount=4)

set_border(ws_report, 'A1:C3')
ws_report.merge_cells('A1:C3')
ws_report.merge_cells('A4:B4')

set_border(ws_report, 'A22:B22')
ws_report.merge_cells('A22:B22')

thin = Side(border_style="thin", color="00000000")

a1 = ws_report['A1']
a1.value = 'RELATÓRIO DE INVENTÁRIO'
a1.font = Font(b=True, size=20)
a1.alignment = Alignment(horizontal="center", vertical="center")

a4 = ws_report['A4']
a4.value = 'Data do relatório:'
a4.alignment = Alignment(horizontal="right")
a4.border = Border(left=thin)

set_font_bold(ws_report, 'A5:C5')
set_align_center(ws_report, 'A5:C5')
set_align_center(ws_report, 'A6:A21')
set_style_currency(ws_report, 'C6:C22')
set_font_size(ws_report, 'C6:C21')
set_style_color_yellow(ws_report, 'C6:C18')
set_border(ws_report, 'A5:C21')

c4 = ws_report['C4']
c4.value = report_day
c4.alignment = Alignment(horizontal='center')
c4.border = Border(right=thin)

a22 = ws_report['A22']
a22.value = 'TOTAL GERAL:'
a22.font = Font(b=True)
a22.alignment = Alignment(horizontal="right")

c22 = ws_report['C22']
c22.value = sum_total
c22.font = Font(b=True)
c22.border = Border(left=thin, right=thin, bottom=thin, top=thin)

b24 = ws_report['B24']
b24.value = 'TOTAL CONTÁBIL:'
b24.font = Font(b=True)
b24.alignment = Alignment(horizontal='right')
b24.fill = PatternFill('solid', fgColor='00FFFF99')
b24.border = Border(top=thin, bottom=thin, left=thin, right=thin)

c24 = ws_report['C24']
c24.value = sum_contabil
c24.style = 'Currency'
c24.font = Font(b=True)
c24.fill = PatternFill('solid', fgColor='00FFFF99')
c24.border = Border(top=thin, bottom=thin, left=thin, right=thin)

wb_report.save(WB_REPORT_PATH)
conn.close()
