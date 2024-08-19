from datetime import datetime
from pytz import timezone
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.worksheet.cell_range import CellRange


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")

    rng = CellRange(cell_range)
    for row in ws.iter_rows(min_row=rng.min_row,
                            min_col=rng.min_col,
                            max_col=rng.max_col,
                            max_row=rng.max_row):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def set_font_bold(ws, cell_range):
    rng = CellRange(cell_range)
    for row in ws.iter_rows(min_row=rng.min_row,
                            min_col=rng.min_col,
                            max_col=rng.max_col,
                            max_row=rng.max_row):
        for cell in row:
            cell.font = Font(b=True)

def set_font_size(ws, cell_range):
    rng = CellRange(cell_range)
    for row in ws.iter_rows(min_row=rng.min_row,
                            min_col=rng.min_col,
                            max_col=rng.max_col,
                            max_row=rng.max_row):
        for cell in row:
            cell.font = Font(size=11)

def set_align_center(ws, cell_range):
    rng = CellRange(cell_range)
    for row in ws.iter_rows(min_row=rng.min_row,
                            min_col=rng.min_col,
                            max_col=rng.max_col,
                            max_row=rng.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

def set_style_currency(ws, cell_range):
    rng = CellRange(cell_range)
    for row in ws.iter_rows(min_row=rng.min_row,
                            min_col=rng.min_col,
                            max_col=rng.max_col,
                            max_row=rng.max_row):
        for cell in row:
            cell.style = 'Currency'

def set_style_color_yellow(ws, cell_range):
    rng = CellRange(cell_range)
    for row in ws.iter_rows(min_row=rng.min_row,
                            min_col=rng.min_col,
                            max_col=rng.max_col,
                            max_row=rng.max_row):
        for cell in row:
            cell.fill = PatternFill('solid', fgColor='00FFFF99')

ROOT_FOLDER = Path(__file__).parent
WB_PRICE_PATH = ROOT_FOLDER / 'InventItemPrice_Consulta.xlsx'
WB_INVENT_PATH = ROOT_FOLDER / 'Estoque.xlsx'
WB_REPORT_PATH = ROOT_FOLDER / 'Relatório de Estoque.xlsx'
dep01, dep02, dep03, dep04, dep11, dep12, dep13, dep14, dep15 = 0, 0, 0, 0, 0, 0, 0, 0, 0
dep19, dep20, dep21, dep96, dep97, dep98, dep00 = 0, 0, 0, 0, 0, 0, 0
sum_total = 0
day = datetime.now(timezone('America/Sao_Paulo'))
fmt = '%d/%m/%Y'
report_day = datetime.strftime(day, fmt)

data = [
    ['Nº do item', 'Nome do item', 'Depósito', 'Localização', 'Estoque físico',
      'Unidade de estoque', 'Preço unitário', 'Preço total'],
]

invent_sum = [
    ['Depósito', 'Nome', 'Valor'],
    ['1', 'Almoxarifado', dep01],
    ['2', 'Material usado', dep02],
    ['3', 'Produto acabado', dep03],
    ['4', 'Em produção', dep04],
    ['11', 'Remessa conserto', dep11],
    ['12', 'Remessa garantia', dep12],
    ['13', 'Demonstração Peças', dep13],
    ['14', 'Demonstraçaõ Máquinas', dep14],
    ['15', 'Remessa industrialização', dep15],
    ['19', 'Conserto máquinas terceiros', dep19],
    ['20', 'Assistência técnica', dep20],
    ['21', 'Ferramentas', dep21],
    ['96', 'Descarte assistência técnica', dep96],
    ['97', 'Descarte comercial', dep97],
    ['98', 'Descarte produção', dep98],
    [' ', 'Sem depósito', dep00],
]

# Carregando um arquivo do Excel
wb_price: Workbook = load_workbook(WB_PRICE_PATH, data_only=True)
wb_invent: Workbook = load_workbook(WB_INVENT_PATH, data_only=True)
wb_report = Workbook()

# Nome para a planilha
sheet_name_price = 'Consulta1'
sheet_name_invent = 'Planilha1'
sheet_name_report = 'Estoque'
sheet_name_report_sum = 'Inventário'

wb_report.create_sheet(sheet_name_report, 0)
wb_report.create_sheet(sheet_name_report_sum, 1)

# Selecionou a planilha
ws_price: Worksheet = wb_price[sheet_name_price]
ws_invent: Worksheet = wb_invent[sheet_name_invent]
ws_report: Worksheet = wb_report[sheet_name_report]

# Remover uma planilha
wb_report.remove(wb_report['Sheet'])

row_price: tuple[Cell]
row_invent: tuple[Cell]
for row_invent in ws_invent.iter_rows(min_row=2, max_col=5):
    item, name, dep, loc, fis_inv = row_invent
    if fis_inv.value is not None:
        for row_price in ws_price.iter_rows(min_row=2, max_col=3):
            itemid, price, unitid = row_price
            if item.value == itemid.value:
                total = price.value * fis_inv.value
                sum_total += total
                data += [[item.value, name.value, dep.value, loc.value, fis_inv.value, unitid.value, price.value, total]]
                if dep.value == '01':
                    dep01 += total
                elif dep.value == '02':
                    dep02 += total
                elif dep.value == '03':
                    dep03 += total
                elif dep.value == '04':
                    dep04 += total
                elif dep.value == '11':
                    dep11 += total
                elif dep.value == '12':
                    dep12 += total
                elif dep.value == '13':
                    dep13 += total
                elif dep.value == '14':
                    dep14 += total
                elif dep.value == '15':
                    dep15 += total
                elif dep.value == '19':
                    dep19 += total
                elif dep.value == '20':
                    dep20 += total
                elif dep.value == '21':
                    dep21 += total
                elif dep.value == '96':
                    dep96 += total
                elif dep.value == '97':
                    dep97 += total
                elif dep.value == '98':
                    dep98 += total
                elif dep.value == None:
                    dep00 += total

    sum_contabil = dep00+dep01+dep02+dep03+dep04+dep11+dep12+dep13+dep14+dep15+dep19+dep20+dep21

for line in data:
    ws_report.append(line)


invent_sum = [
    ['Depósito', 'Nome', 'Valor'],
    [' ', 'Sem depósito', dep00],
    ['1', 'Almoxarifado', dep01],
    ['2', 'Material usado', dep02],
    ['3', 'Produto acabado', dep03],
    ['4', 'Em produção', dep04],
    ['11', 'Remessa conserto', dep11],
    ['12', 'Remessa garantia', dep12],
    ['13', 'Demonstração Peças', dep13],
    ['14', 'Demonstraçaõ Máquinas', dep14],
    ['15', 'Remessa industrialização', dep15],
    ['19', 'Conserto máquinas terceiros', dep19],
    ['20', 'Assistência técnica', dep20],
    ['21', 'Ferramentas', dep21],
    ['96', 'Descarte assistência técnica', dep96],
    ['97', 'Descarte comercial', dep97],
    ['98', 'Descarte produção', dep98],
]

ws_report: Worksheet = wb_report[sheet_name_report_sum]

for line in invent_sum:
    ws_report.append(line)

ws_report.insert_rows(1, amount=4)

set_border(ws_report, 'A1:C3')
ws_report.merge_cells('A1:C3')
ws_report.merge_cells('A4:B4')

set_border(ws_report, 'A22:B22')
ws_report.merge_cells('A22:B22')

thin = Side(border_style="thin", color="00000000")

a1 = ws_report['A1']
a1.value = 'RELATÓRIO DE INVENTÁRIO'
a1.font = Font(b=True, size=20)
a1.alignment = Alignment(horizontal="center", vertical="center")

a4 = ws_report['A4']
a4.value = 'Data do relatório:'
a4.alignment = Alignment(horizontal="right")
a4.border = Border(left=thin)

set_font_bold(ws_report, 'A5:C5')
set_align_center(ws_report, 'A5:C5')
set_align_center(ws_report, 'A6:A21')
set_style_currency(ws_report, 'C6:C22')
set_font_size(ws_report, 'C6:C21')
set_style_color_yellow(ws_report, 'C6:C18')
set_border(ws_report, 'A5:C21')

c4 = ws_report['C4']
c4.value = report_day
c4.alignment = Alignment(horizontal='center')
c4.border = Border(right=thin)

a22 = ws_report['A22']
a22.value = 'TOTAL GERAL:'
a22.font = Font(b=True)
a22.alignment = Alignment(horizontal="right")

c22 = ws_report['C22']
c22.value = sum_total
c22.font = Font(b=True)
c22.border = Border(left=thin, right=thin, bottom=thin, top=thin)

b24 = ws_report['B24']
b24.value = 'TOTAL CONTÁBIL:'
b24.font = Font(b=True)
b24.alignment = Alignment(horizontal='right')
b24.fill = PatternFill('solid', fgColor='00FFFF99')
b24.border = Border(top=thin, bottom=thin, left=thin, right=thin)

c24 = ws_report['C24']
c24.value = sum_contabil
c24.style = 'Currency'
c24.font = Font(b=True)
c24.fill = PatternFill('solid', fgColor='00FFFF99')
c24.border = Border(top=thin, bottom=thin, left=thin, right=thin)

wb_report.save(WB_REPORT_PATH)
